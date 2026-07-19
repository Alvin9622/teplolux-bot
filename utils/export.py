"""Shaxsiy samaradorlik hisobotini Excel qilib eksport."""
import datetime
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import database as db


async def build_excel_report(user_id: int) -> bytes:
    stats = await db.get_focus_stats(user_id)
    since_30 = (datetime.datetime.now() - datetime.timedelta(days=30)).strftime("%Y-%m-%d")
    logs  = await db.get_time_logs_detail(user_id, since_30)
    goals = await db.get_goals(user_id, active_only=False)

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E86DE")

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # 1) Umumiy
    ws1 = wb.active
    ws1.title = "Umumiy"
    ws1.append(["Ko'rsatkich", "Qiymat"])
    style_header(ws1, 2)
    ws1.append(["Joriy streak (kun)", stats["current_streak"]])
    ws1.append(["Eng uzun streak (kun)", stats["longest_streak"]])
    ws1.append(["Jami fokus (soat)", round(stats["total_focus_minutes"] / 60, 1)])
    ws1.append(["Jami pomodorolar", stats["total_pomodoros"]])
    ws1.append(["Fokus ballari", stats["focus_points"]])
    ws1.append(["Hisobot sanasi", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws1.column_dimensions["A"].width = 26
    ws1.column_dimensions["B"].width = 22

    # 2) Vaqt loglari (30 kun)
    ws2 = wb.create_sheet("Vaqt (30 kun)")
    ws2.append(["Sana", "Kategoriya", "Ish", "Daqiqa"])
    style_header(ws2, 4)
    for l in logs:
        ws2.append([l["log_date"], l["category"], l["task_name"],
                    round(l["duration_seconds"] / 60)])
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 42
    ws2.column_dimensions["D"].width = 10

    # 3) Maqsadlar
    ws3 = wb.create_sheet("Maqsadlar")
    ws3.append(["Maqsad", "Tur", "Joriy", "Maqsad", "Birlik", "%", "Holat"])
    style_header(ws3, 7)
    for g in goals:
        current = await db.compute_goal_current(g)
        pct = round(current / g["target_value"] * 100) if g["target_value"] else 0
        ws3.append([g["title"], g["kind"], current, g["target_value"],
                    g["unit"], min(pct, 100), g["status"]])
    ws3.column_dimensions["A"].width = 34
    ws3.column_dimensions["G"].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
