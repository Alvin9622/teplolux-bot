import io
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


STATUS_UZ = {
    "new": "Yangi", "in_progress": "Jarayonda",
    "done": "Bajarildi", "cancelled": "Bekor", "review": "Tekshiruvda"
}
PRIORITY_UZ = {"high": "Yuqori", "medium": "O'rta", "low": "Past"}


def _fmt(iso):
    if not iso:
        return ""
    try:
        return datetime.date.fromisoformat(iso[:10]).strftime("%d.%m.%Y")
    except Exception:
        return iso


def build_tasks_excel(tasks, month=None, year=None):
    wb = openpyxl.Workbook()
    ws = wb.active

    now = datetime.datetime.now()
    if month and year:
        months_uz = ["Yanvar","Fevral","Mart","Aprel","May","Iyun",
                     "Iyul","Avgust","Sentabr","Oktabr","Noyabr","Dekabr"]
        ws.title = f"{months_uz[month-1]} {year}"
        title = f"Teplolux — Vazifalar hisoboti: {months_uz[month-1]} {year}"
    else:
        ws.title = "Barcha vazifalar"
        title = f"Teplolux — Barcha vazifalar ({now.strftime('%d.%m.%Y')})"

    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    done_fill    = PatternFill("solid", fgColor="E2EFDA")
    overdue_fill = PatternFill("solid", fgColor="FFE0E0")
    cancel_fill  = PatternFill("solid", fgColor="F2F2F2")
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.merge_cells("A1:K1")
    ws["A1"] = title
    ws["A1"].font      = Font(bold=True, size=13, color="1F4E79")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:K2")
    ws["A2"] = f"Yaratilgan: {now.strftime('%d.%m.%Y %H:%M')}  |  Jami: {len(tasks)} ta"
    ws["A2"].font      = Font(italic=True, size=10, color="666666")
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 18

    headers = ["#", "Vazifa nomi", "Mas'ul hodim", "Lavozim",
               "Kategoriya", "Holat", "Ustuvorlik",
               "Muddat", "Yaratilgan", "Bajarildi %", "Yaratgan"]
    widths  = [5, 35, 22, 20, 18, 14, 12, 14, 14, 12, 18]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 22

    today = datetime.date.today().isoformat()

    for row_idx, t in enumerate(tasks, 4):
        is_done    = t["status"] == "done"
        is_cancel  = t["status"] == "cancelled"
        is_overdue = (t.get("deadline") and t["deadline"] < today
                      and t["status"] not in ("done", "cancelled"))

        row_fill = done_fill if is_done else (cancel_fill if is_cancel else
                   (overdue_fill if is_overdue else None))

        values = [
            row_idx - 3,
            t.get("title") or "",
            t.get("assignee_name") or "—",
            t.get("assignee_position") or "—",
            t.get("category") or "—",
            STATUS_UZ.get(t.get("status"), t.get("status") or "—"),
            PRIORITY_UZ.get(t.get("priority"), t.get("priority") or "—"),
            _fmt(t.get("deadline")),
            _fmt(t.get("created_at")),
            t.get("progress_pct") or 0,
            t.get("creator_name") or "—",
        ]
        aligns = [center, left, left, left, left, center, center,
                  center, center, center, left]

        for col, (val, aln) in enumerate(zip(values, aligns), 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = aln
            cell.border    = border
            if row_fill:
                cell.fill = row_fill

        ws.row_dimensions[row_idx].height = 20

    # Summary block
    summary_row = len(tasks) + 5
    done_count  = sum(1 for t in tasks if t["status"] == "done")
    prog_count  = sum(1 for t in tasks if t["status"] == "in_progress")
    over_count  = sum(1 for t in tasks if (t.get("deadline") and t["deadline"] < today
                      and t["status"] not in ("done","cancelled")))
    pct = round(done_count / len(tasks) * 100) if tasks else 0

    ws.merge_cells(f"A{summary_row}:C{summary_row}")
    ws[f"A{summary_row}"] = "📊 JAMI STATISTIKA"
    ws[f"A{summary_row}"].font      = Font(bold=True, size=11, color="1F4E79")
    ws[f"A{summary_row}"].alignment = center

    for r, (label, val) in enumerate([
        ("Jami vazifalar", len(tasks)),
        ("Bajarildi ✅",   done_count),
        ("Jarayonda 🔄",   prog_count),
        ("Kechikkan 🔴",   over_count),
        ("Bajarilish %",   f"{pct}%"),
    ], summary_row + 1):
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)

    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
